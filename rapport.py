#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GaraServer
Copyright 2016 Nicola Ferruzzi <nicola.ferruzzi@gmail.com>
License: GPLv3 (see LICENSE)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, Alignment
from openpyxl.styles import PatternFill, Border, Side, Protection

def baseGen():
    return [
        ('N° pett.', lambda n, user_values, user_info: n),
        ('Concorrente', lambda n, user_values, user_info: user_info['nickname'] or '')
    ]


def dumpRows(gara, connection, worksheet, generator, conf, resuls_required=True, trial_required=None):
    alignment = Alignment(horizontal='center',
                          vertical='bottom',
                          wrap_text=True)

    aleft = Alignment(horizontal='left',
                      vertical='bottom',
                      wrap_text=True)

    thin_border = Border(left=Side(style='hair'),
                         right=Side(style='hair'),
                         top=Side(style='hair'),
                         bottom=Side(style='hair'))

    row = 3
    for x, g in enumerate(generator):
        title = g[0]
        c = worksheet.cell(column=x+1, row=row, value=title)
        c.font = Font(bold=True)
        if x != 1:
            c.alignment = alignment
        else:
            c.alignment = aleft
        c.border = thin_border

    row += 1
    users_with_a_vote = gara.getAllUsersWithAVote(connection)
    for user in users_with_a_vote:
        user_values = gara.getUser(connection, user)
        results = user_values.get('results')
        if resuls_required and results is None:
            continue
        if trial_required is not None:
            votes = user_values['trials'][trial_required]['votes'].values()
            s = set(votes)
            if None in s and len(s) == 1:
                continue
        user_info = gara.getUserInfo(connection, user)
        for x, g in enumerate(generator):
            gen = g[1]
            val = gen(user, user_values, user_info)
            v = worksheet.cell(column=x+1, row=row+1, value=val)

            if isinstance(val, float):
                v.number_format = '0.00'
            if x != 1:
                v.alignment = alignment
            else:
                v.alignment = aleft

            v.border = thin_border
            if x == 0:
                v.font = Font(bold=True)

        row += 1

    worksheet.column_dimensions['A'].width = 8
    worksheet.column_dimensions['B'].width = 20
    for x in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'L']:
        worksheet.column_dimensions[x].width = 10
    row = worksheet.row_dimensions[1].height = 80
    row = worksheet.row_dimensions[3].height = 60

    worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToHeight = 0
    worksheet.page_setup.fitToWidth = 1
    worksheet.print_options.horizontalCentered = True
    worksheet.print_options.verticalCentered = False
    worksheet.header_footer.center_footer.text = "Pagina &P/&N"
    worksheet.header_footer.center_footer.font_size = 8
    #worksheet.freeze_panes = worksheet.cell('L3')

    worksheet.merge_cells(start_row=1,
                          end_row=2,
                          start_column=1,
                          end_column=len(generator))

    worksheet['A1'] = conf['description'] + '\n' + conf['date'].strftime('%d-%m-%Y')+'\n\n'+worksheet.title+'\n'
    worksheet['A1'].alignment = alignment
    worksheet['A1'].font = Font(size=16, bold=True)

    worksheet.cell(row=3, column=2).border = thin_border
    worksheet.add_print_title(4, rows_or_cols='rows')


def generateRapport(gara, connection, filename, include):
    conf = gara.getConfiguration(connection)
    workbook = Workbook()

    generator = baseGen()
    for t in range(0, conf['nTrials']):
        v = ('Punteggio prova {}'.format(t+1), lambda n, user_values, user_info, trial=t: user_values['trials'][trial]['score_bonus'] or 0.0)
        generator.append(v)

    generator.append(('Media punteggi', lambda n, user_values, user_info: user_values['results']['average']))
    generator.append(('Media punteggi con crediti', lambda n, user_values, user_info: user_values['results']['average_bonus']))
    generator.append(('Somma punteggi prove con crediti', lambda n, user_values, user_info: user_values['results']['sum']))

    worksheet = workbook.active
    worksheet.title = 'Risultati'
    dumpRows(gara, connection, worksheet, generator, conf)

    for t in range(conf['nTrials']):
        worksheet = workbook.create_sheet(title="Prova {}".format(t+1))
        generator = baseGen()

        for j in range(1, conf['nJudges']+1):
            # Dump values as is
            def aritmetica(n, user_values, user_info, trial=t, judge=j):
                return user_values['trials'][trial]['votes'][judge]

            # Remove first min and first max
            def mediata(n, user_values, user_info, trial=t, judge=j):
                v = dict(user_values['trials'][trial]['votes'])
                imax = max(v, key=v.get)
                imin = min(v, key=v.get)
                v[imax] = '-'
                v[imin] = '-'
                return v[judge]

            v = ('Giudice {}'.format(j), aritmetica if (conf['average'] == 0 or include) else mediata)
            generator.append(v)

        generator.append(('Punteggio', lambda n, user_values, user_info, trial=t: user_values['trials'][trial]['score'] or 0.0))
        generator.append(('Punteggio con crediti', lambda n, user_values, user_info, trial=t: user_values['trials'][trial]['score_bonus'] or 0.0))
        if t != 0:
            generator.append(('Media punteggi prove con crediti',
                             lambda n, user_values, user_info, trial=t: user_values['trials'][trial]['average_bonus'] or 0.0))

        dumpRows(gara, connection, worksheet, generator, conf, False, trial_required=t)

    workbook.save(filename)
